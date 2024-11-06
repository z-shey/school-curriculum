import os
import pandas as pd
from openpyxl import load_workbook

# 指定目录路径
directory = '2024-09-07_00-31-51'  # 目标目录
target_file = 'result1_1-5.xlsx'  # 目标文件
target_book = load_workbook(target_file)  # 目标文件

# # 遍历目录下的所有文件
for filename in os.listdir(directory):
    # 解析文件名获取年份和季度序号
    parts = filename.split('_')  # 分隔符
    year = parts[0]  # 年份
    season = int(parts[2].split('.')[0])  # 季度

    # 加载源文件
    source_book = load_workbook(directory + "/" + filename)

    # 根据年份选定工作簿
    target_sheet = target_book[str(year)]

    # 解析源数据范围
    if season == 1:
        # 源文件从第二行开始复制（第一行为表头）
        # 目标文件从第二行第三列开始粘贴
        start_row = 2
        for row in source_book.active.iter_rows(min_row=start_row, max_col=source_book.active.max_column,
                                                max_row=source_book.active.max_row):
            for cell in row:
                target_sheet.cell(row=cell.row + start_row - 1, column=3, value=cell.value)  # 从第三列开始粘贴
        # pass
    if season == 2:
        # 源文件从第二行开始复制（第一行为表头）
        # 目标文件从C56开始粘贴
        pass

    # 复制数据

# 保存目标Excel文件
target_book.save(target_file)
